from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Import sheet
wb = load_workbook('data.xlsx')
sheet = wb['dat']


# Set the coordinates of the cells
cords = [wb.active.min_row,
         wb.active.min_column,
         wb.active.max_row,
         wb.active.max_column]

# Make the chart data
data = Reference(sheet, min_row=cords[0], min_col=cords[1]+1, max_row=cords[2], max_col=cords[3])
catagories = Reference(sheet, min_row=cords[0]+1, min_col=cords[1], max_row=cords[2], max_col=cords[3]-1)

# Make the Barchart
barchart = BarChart()
barchart.add_data(data, titles_from_data=1)
barchart.set_categories(catagories)

sheet.add_chart(barchart, "G3")

barchart.title = "Total GDP"
barchart.style = 3
wb.save('barchart.xlsx')

# Set the coordinates of the cells
cords = [wb.active.min_row,
         wb.active.min_column,
         wb.active.max_row,
         wb.active.max_column]

# Make the chart data
data = Reference(sheet, min_row=cords[0], min_col=cords[1]+1, max_row=cords[2], max_col=cords[3])
catagories = Reference(sheet, min_row=cords[0]+1, min_col=cords[1], max_row=cords[2], max_col=cords[3]-1)

# Make the Barchart
barchart = BarChart()
barchart.add_data(data, titles_from_data=1)
barchart.set_categories(catagories)

sheet.add_chart(barchart, "G3")

barchart.title = "Total GDP"
barchart.style = 3
wb.save('barchart.xlsx')


